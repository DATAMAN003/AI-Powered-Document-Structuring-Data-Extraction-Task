"""
AI-Powered Document Extraction - Intelligent Version
Uses Groq API for generalized extraction from ANY PDF document
"""

import os
import re
import json
import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, List, Any
from groq import Groq


class AIDocumentExtractor:
    """Intelligent document extractor using Groq AI for any PDF type"""
    
    def __init__(self, pdf_path: str, groq_api_key: str = None):
        self.pdf_path = pdf_path
        self.raw_text = ""
        self.structured_data = []
        
        # Initialize Groq client
        api_key = groq_api_key or os.environ.get("GROQ_API_KEY")
        if not api_key:
            raise ValueError("Groq API key required. Set GROQ_API_KEY environment variable or pass as parameter.")
        
        self.client = Groq(api_key=api_key)
        
    def extract_text_from_pdf(self) -> str:
        """Extract all text content from PDF"""
        with open(self.pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text()
        self.raw_text = text
        return text
    
    def analyze_document_with_ai(self) -> List[Dict[str, Any]]:
        """
        Use Groq AI to intelligently analyze and extract structured data
        Works with ANY type of document
        """
        print("\n🤖 Using AI to analyze document...")
        
        # Step 1: Identify document type and structure
        doc_type = self._identify_document_type()
        print(f"✓ Document type identified: {doc_type}")
        
        # Step 2: Extract structured data based on document type
        structured_data = self._extract_structured_data(doc_type)
        print(f"✓ Extracted {len(structured_data)} data entries")
        
        self.structured_data = structured_data
        return structured_data
    
    def _identify_document_type(self) -> str:
        """Use AI to identify the type of document"""
        prompt = f"""Analyze this document text and identify its type (e.g., resume, invoice, contract, report, personal profile, etc.).

Document text:
{self.raw_text[:2000]}...

Respond with ONLY the document type in 2-3 words. Examples: "Personal Resume", "Sales Invoice", "Legal Contract", "Technical Report"."""

        response = self.client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
            max_tokens=50
        )
        
        return response.choices[0].message.content.strip()
    
    def _extract_structured_data(self, doc_type: str) -> List[Dict[str, Any]]:
        """Use AI to extract structured key-value pairs from document"""
        
        # Split text into chunks if too long
        chunks = self._split_text_into_chunks(self.raw_text, max_length=6000)
        all_data = []
        
        for i, chunk in enumerate(chunks):
            print(f"  Processing chunk {i+1}/{len(chunks)}...")
            
            prompt = f"""You are an expert data extraction system. Extract ALL key information from this {doc_type} document.

Document text:
{chunk}

Instructions:
1. Identify ALL important information (names, dates, numbers, facts, etc.)
2. Organize into logical categories
3. Create key-value pairs for each piece of information
4. Add a brief comment explaining the significance of each data point
5. Preserve original wording - do NOT paraphrase

Return a JSON array with this EXACT structure (use these exact key names):
[
  {{
    "Category": "Category Name",
    "Key": "Field Name",
    "Value": "Extracted Value",
    "Comments": "Brief explanation of significance"
  }}
]

IMPORTANT: Use "Category", "Key", "Value", "Comments" (with capital letters).
Extract EVERYTHING - leave nothing out. Be thorough and comprehensive."""

            response = self.client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.2,
                max_tokens=4000
            )
            
            # Parse JSON response
            try:
                content = response.choices[0].message.content.strip()
                # Extract JSON from markdown code blocks if present
                if "```json" in content:
                    content = content.split("```json")[1].split("```")[0].strip()
                elif "```" in content:
                    content = content.split("```")[1].split("```")[0].strip()
                
                chunk_data = json.loads(content)
                
                # Normalize keys to match Excel export format
                normalized_data = []
                for item in chunk_data:
                    normalized_item = {
                        'Category': item.get('Category') or item.get('category', 'Uncategorized'),
                        'Key': item.get('Key') or item.get('key', 'Unknown'),
                        'Value': item.get('Value') or item.get('value', ''),
                        'Comments': item.get('Comments') or item.get('comment') or item.get('comments', '')
                    }
                    normalized_data.append(normalized_item)
                
                all_data.extend(normalized_data)
                print(f"    ✓ Extracted {len(normalized_data)} entries from chunk {i+1}")
                
            except json.JSONDecodeError as e:
                print(f"  Warning: Could not parse AI response for chunk {i+1}: {e}")
                print(f"  Response was: {content[:200]}...")
                # Fallback: try to extract data manually
                fallback_data = self._fallback_extraction(chunk)
                all_data.extend(fallback_data)
        
        # Remove duplicates
        all_data = self._remove_duplicates(all_data)
        
        return all_data
    
    def _split_text_into_chunks(self, text: str, max_length: int = 6000) -> List[str]:
        """Split text into manageable chunks for AI processing"""
        if len(text) <= max_length:
            return [text]
        
        chunks = []
        sentences = text.split('. ')
        current_chunk = ""
        
        for sentence in sentences:
            if len(current_chunk) + len(sentence) < max_length:
                current_chunk += sentence + ". "
            else:
                if current_chunk:
                    chunks.append(current_chunk.strip())
                current_chunk = sentence + ". "
        
        if current_chunk:
            chunks.append(current_chunk.strip())
        
        return chunks
    
    def _fallback_extraction(self, text: str) -> List[Dict[str, Any]]:
        """Fallback extraction using regex patterns if AI parsing fails"""
        data = []
        
        # Extract dates
        dates = re.findall(r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b|\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]* \d{1,2},? \d{4}\b', text)
        for date in dates[:5]:  # Limit to first 5
            data.append({
                "Category": "Dates",
                "Key": "Date Found",
                "Value": date,
                "Comments": "Date extracted from document"
            })
        
        # Extract numbers with context
        numbers = re.findall(r'\b\d+(?:,\d{3})*(?:\.\d+)?\b', text)
        for num in numbers[:10]:  # Limit to first 10
            data.append({
                "Category": "Numerical Data",
                "Key": "Number Found",
                "Value": num,
                "Comments": "Numerical value from document"
            })
        
        return data
    
    def _remove_duplicates(self, data: List[Dict]) -> List[Dict]:
        """Remove duplicate entries based on key and value"""
        seen = set()
        unique_data = []
        
        for entry in data:
            # Use both lowercase and uppercase keys for compatibility
            key = entry.get('Key') or entry.get('key', '')
            value = entry.get('Value') or entry.get('value', '')
            key_value = (key, value)
            if key_value not in seen and key and value:
                seen.add(key_value)
                unique_data.append(entry)
        
        return unique_data
    
    def export_to_excel(self, output_path: str):
        """Export structured data to Excel with professional formatting"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        
        # Define headers
        headers = ['Category', 'Key', 'Value', 'Comments']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Add data rows
        for entry in self.structured_data:
            ws.append([
                entry.get('category', ''),
                entry.get('key', ''),
                entry.get('value', ''),
                entry.get('comment', '')
            ])
        
        # Apply styling
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 70
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save workbook
        wb.save(output_path)
        print(f"✓ Excel file created: {output_path}")
        print(f"✓ Total entries extracted: {len(self.structured_data)}")


def main():
    """Main execution function"""
    print("=" * 80)
    print("AI-Powered Document Extraction - Intelligent Version")
    print("Powered by Groq AI - Works with ANY PDF document")
    print("=" * 80)
    
    # Check for API key
    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        print("\n❌ Error: GROQ_API_KEY environment variable not set")
        print("\nTo set it:")
        print("  Windows: set GROQ_API_KEY=your_api_key_here")
        print("  Mac/Linux: export GROQ_API_KEY=your_api_key_here")
        print("\nGet your free API key at: https://console.groq.com")
        return
    
    # Initialize extractor
    pdf_file = "Data Input.pdf"
    if not os.path.exists(pdf_file):
        print(f"\n❌ Error: {pdf_file} not found")
        return
    
    extractor = AIDocumentExtractor(pdf_file)
    
    # Extract text
    print("\n[1/3] Extracting text from PDF...")
    text = extractor.extract_text_from_pdf()
    print(f"✓ Extracted {len(text)} characters")
    
    # AI Analysis
    print("\n[2/3] AI analyzing document...")
    data = extractor.analyze_document_with_ai()
    
    # Export to Excel
    print("\n[3/3] Exporting to Excel...")
    extractor.export_to_excel("Output_AI.xlsx")
    
    # Summary
    print("\n" + "=" * 80)
    print("EXTRACTION SUMMARY")
    print("=" * 80)
    categories = {}
    for entry in data:
        cat = entry.get('category', 'Uncategorized')
        categories[cat] = categories.get(cat, 0) + 1
    
    for category, count in sorted(categories.items()):
        print(f"  • {category}: {count} entries")
    
    print(f"\n  TOTAL: {len(data)} data points extracted")
    print("\n✓ Process completed successfully!")
    print("✓ AI-powered extraction works with ANY PDF document type!")
    print("=" * 80)


if __name__ == "__main__":
    main()
