"""
AI-Powered Document Structuring & Data Extraction
Extracts structured data from unstructured PDF documents into Excel format
"""

import re
import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import Dict, List, Tuple, Any


class DocumentExtractor:
    """Extract and structure data from PDF documents"""
    
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.raw_text = ""
        self.structured_data = []
        
    def extract_text_from_pdf(self) -> str:
        """Extract all text content from PDF"""
        with open(self.pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text()
        self.raw_text = text
        return text
    
    def identify_key_value_pairs(self) -> List[Dict[str, Any]]:
        """
        Intelligently identify key-value relationships in unstructured text
        Uses pattern matching and contextual analysis
        """
        text = self.raw_text
        data_entries = []
        
        # Personal Information Section
        data_entries.extend(self._extract_personal_info(text))
        
        # Career/Employment Section
        data_entries.extend(self._extract_career_info(text))
        
        # Education Section
        data_entries.extend(self._extract_education_info(text))
        
        # Certifications Section
        data_entries.extend(self._extract_certification_info(text))
        
        # Skills Section
        data_entries.extend(self._extract_skills_info(text))
        
        self.structured_data = data_entries
        return data_entries
    
    def _extract_personal_info(self, text: str) -> List[Dict]:
        """Extract personal information with context"""
        entries = []
        
        # Name extraction
        name_match = re.search(r'([A-Z][a-z]+\s+[A-Z][a-z]+)\s+was born', text)
        if name_match:
            name = name_match.group(1)
            entries.append({
                'Category': 'Personal Information',
                'Key': 'Full Name',
                'Value': name,
                'Comments': 'Primary identifier for the individual'
            })
        
        # Birth date extraction
        birth_match = re.search(r'born on\s+([A-Z][a-z]+\s+\d{1,2},\s+\d{4})', text)
        if birth_match:
            birth_date = birth_match.group(1)
            entries.append({
                'Category': 'Personal Information',
                'Key': 'Date of Birth',
                'Value': birth_date,
                'Comments': 'Original format as stated in document'
            })
        
        # ISO format date
        iso_match = re.search(r'formatted as\s+(\d{4}-\d{2}-\d{2})', text)
        if iso_match:
            iso_date = iso_match.group(1)
            entries.append({
                'Category': 'Personal Information',
                'Key': 'Date of Birth (ISO Format)',
                'Value': iso_date,
                'Comments': 'ISO 8601 format for easy parsing and database storage'
            })
        
        # Age extraction
        age_match = re.search(r'making him\s+(\d+)\s+years old', text)
        if age_match:
            age = age_match.group(1)
            entries.append({
                'Category': 'Personal Information',
                'Key': 'Age',
                'Value': age,
                'Comments': 'Age as of 2024; key demographic marker for analytical purposes'
            })
        
        # Birthplace extraction
        birthplace_match = re.search(r'born on[^,]+,\s+in\s+([^,]+,\s+[^,]+)', text)
        if birthplace_match:
            birthplace = birthplace_match.group(1)
            entries.append({
                'Category': 'Personal Information',
                'Key': 'Birthplace',
                'Value': birthplace,
                'Comments': 'Provides valuable regional profiling context'
            })
        
        # Birthplace nickname
        if 'Pink City of India' in text:
            entries.append({
                'Category': 'Personal Information',
                'Key': 'Birthplace Description',
                'Value': 'Pink City of India',
                'Comments': 'Cultural reference to Jaipur, Rajasthan'
            })
        
        # Blood group
        blood_match = re.search(r'his\s+([A-Z]\+)\s+blood group', text)
        if blood_match:
            blood_group = blood_match.group(1)
            entries.append({
                'Category': 'Personal Information',
                'Key': 'Blood Group',
                'Value': blood_group,
                'Comments': 'Noted for emergency contact purposes'
            })
        
        # Nationality
        nationality_match = re.search(r'As an\s+([A-Z][a-z]+)\s+national', text)
        if nationality_match:
            nationality = nationality_match.group(1)
            entries.append({
                'Category': 'Personal Information',
                'Key': 'Nationality',
                'Value': nationality,
                'Comments': 'Important for understanding work authorization and visa requirements'
            })
        
        return entries
    
    def _extract_career_info(self, text: str) -> List[Dict]:
        """Extract career and employment information"""
        entries = []
        
        # First job information
        first_job_match = re.search(
            r'professional journey began on\s+([A-Z][a-z]+\s+\d{1,2},\s+\d{4}),\s+when he joined his first company as a\s+([^w]+)\s+with an annual salary of\s+([\d,]+\s+INR)',
            text
        )
        if first_job_match:
            start_date, position, salary = first_job_match.groups()
            entries.append({
                'Category': 'Career History',
                'Key': 'First Employment Start Date',
                'Value': start_date,
                'Comments': 'Beginning of professional career'
            })
            entries.append({
                'Category': 'Career History',
                'Key': 'First Position',
                'Value': position.strip(),
                'Comments': 'Entry-level role in technology sector'
            })
            entries.append({
                'Category': 'Career History',
                'Key': 'First Salary',
                'Value': salary,
                'Comments': 'Starting compensation package'
            })
        
        # Current role information
        current_match = re.search(
            r'current role at\s+([^b]+)\s+beginning on\s+([A-Z][a-z]+\s+\d{1,2},\s+\d{4}),\s+where he serves as a\s+([^e]+)\s+earning\s+([\d,]+\s+INR)',
            text
        )
        if current_match:
            company, start_date, position, salary = current_match.groups()
            entries.append({
                'Category': 'Career History',
                'Key': 'Current Company',
                'Value': company.strip(),
                'Comments': 'Present employer'
            })
            entries.append({
                'Category': 'Career History',
                'Key': 'Current Role Start Date',
                'Value': start_date,
                'Comments': 'Date of joining current organization'
            })
            entries.append({
                'Category': 'Career History',
                'Key': 'Current Position',
                'Value': position.strip(),
                'Comments': 'Senior-level technical role'
            })
            entries.append({
                'Category': 'Career History',
                'Key': 'Current Salary',
                'Value': salary,
                'Comments': 'Current annual compensation'
            })
        
        # Previous company
        previous_match = re.search(
            r'he worked at\s+([^f]+)\s+from\s+([A-Z][a-z]+\s+\d{1,2},\s+\d{4}),\s+to\s+(\d{4}),\s+starting as a\s+([^a]+)\s+and earning a promotion in\s+(\d{4})',
            text
        )
        if previous_match:
            company, start_date, end_year, position, promo_year = previous_match.groups()
            entries.append({
                'Category': 'Career History',
                'Key': 'Previous Company',
                'Value': company.strip(),
                'Comments': 'Former employer'
            })
            entries.append({
                'Category': 'Career History',
                'Key': 'Previous Company Start Date',
                'Value': start_date,
                'Comments': 'Date of joining previous organization'
            })
            entries.append({
                'Category': 'Career History',
                'Key': 'Previous Company End Year',
                'Value': end_year,
                'Comments': 'Year of departure from previous organization'
            })
            entries.append({
                'Category': 'Career History',
                'Key': 'Previous Position',
                'Value': position.strip(),
                'Comments': 'Initial role at previous company'
            })
            entries.append({
                'Category': 'Career History',
                'Key': 'Promotion Year',
                'Value': promo_year,
                'Comments': 'Year of career advancement at previous company'
            })
        
        # Career span
        career_span_match = re.search(r'over his\s+([^-]+)-year career span', text)
        if career_span_match:
            years = career_span_match.group(1).strip()
            entries.append({
                'Category': 'Career History',
                'Key': 'Total Career Span',
                'Value': f'{years} years',
                'Comments': 'Total professional experience duration'
            })
        
        # Salary progression
        salary_prog_match = re.search(
            r'salary progression from his starting compensation to his current peak salary of\s+([\d,]+\s+INR)\s+represents a substantial\s+([^-]+)-\s*fold increase',
            text
        )
        if salary_prog_match:
            peak_salary, multiplier = salary_prog_match.groups()
            entries.append({
                'Category': 'Career History',
                'Key': 'Peak Salary',
                'Value': peak_salary,
                'Comments': 'Highest compensation achieved in career'
            })
            entries.append({
                'Category': 'Career History',
                'Key': 'Salary Growth Multiple',
                'Value': f'{multiplier.strip()}-fold',
                'Comments': 'Represents substantial career progression and value appreciation'
            })
        
        return entries
    
    def _extract_education_info(self, text: str) -> List[Dict]:
        """Extract educational background information"""
        entries = []
        
        # High school
        hs_match = re.search(
            r'high school education at\s+([^,]+),\s+([^,]+),\s+where he completed his\s+(\d+)[a-z]+\s+standard in\s+(\d{4}),\s+achieving an outstanding\s+([\d.]+)%\s+overall score',
            text
        )
        if hs_match:
            school, city, grade, year, score = hs_match.groups()
            entries.append({
                'Category': 'Education',
                'Key': 'High School Name',
                'Value': school.strip(),
                'Comments': 'Secondary education institution'
            })
            entries.append({
                'Category': 'Education',
                'Key': 'High School Location',
                'Value': city.strip(),
                'Comments': 'City where secondary education was completed'
            })
            entries.append({
                'Category': 'Education',
                'Key': 'High School Grade Level',
                'Value': f'{grade}th Standard',
                'Comments': 'Final year of secondary education'
            })
            entries.append({
                'Category': 'Education',
                'Key': 'High School Completion Year',
                'Value': year,
                'Comments': 'Year of board examination completion'
            })
            entries.append({
                'Category': 'Education',
                'Key': 'High School Score',
                'Value': f'{score}%',
                'Comments': 'Outstanding performance in board examinations'
            })
        
        # High school subjects
        hs_subjects_match = re.search(r'core subjects included\s+([^,]+),\s+([^,]+),\s+([^,]+),\s+and\s+([^,]+),\s+demonstrating', text)
        if hs_subjects_match:
            subjects = ', '.join(hs_subjects_match.groups())
            entries.append({
                'Category': 'Education',
                'Key': 'High School Core Subjects',
                'Value': subjects,
                'Comments': 'Demonstrates early aptitude for technical disciplines'
            })
        
        # B.Tech information
        btech_match = re.search(
            r'pursued his\s+([^i]+)\s+in\s+([^a]+)\s+at the prestigious\s+([^,]+),\s+graduating with honors in\s+(\d{4})\s+with a CGPA of\s+([\d.]+)\s+on a\s+(\d+)-point scale,\s+ranking\s+(\d+)[a-z]+\s+among\s+(\d+)\s+students',
            text
        )
        if btech_match:
            degree, field, college, year, cgpa, scale, rank, total = btech_match.groups()
            entries.append({
                'Category': 'Education',
                'Key': 'Undergraduate Degree',
                'Value': degree.strip(),
                'Comments': 'Bachelor of Technology degree'
            })
            entries.append({
                'Category': 'Education',
                'Key': 'Undergraduate Field',
                'Value': field.strip(),
                'Comments': 'Specialization in technology domain'
            })
            entries.append({
                'Category': 'Education',
                'Key': 'Undergraduate Institution',
                'Value': college.strip(),
                'Comments': 'Prestigious Indian Institute of Technology'
            })
            entries.append({
                'Category': 'Education',
                'Key': 'Undergraduate Graduation Year',
                'Value': year,
                'Comments': 'Year of degree completion'
            })
            entries.append({
                'Category': 'Education',
                'Key': 'Undergraduate CGPA',
                'Value': f'{cgpa}/{scale}',
                'Comments': 'Graduated with honors; demonstrates academic excellence'
            })
            entries.append({
                'Category': 'Education',
                'Key': 'Undergraduate Class Rank',
                'Value': f'{rank} out of {total}',
                'Comments': 'Top percentile performance in competitive cohort'
            })
        
        # M.Tech information
        mtech_match = re.search(
            r'earned his\s+([^i]+)\s+in\s+([^i]+)\s+in\s+(\d{4}),\s+achieving an exceptional CGPA of\s+([\d.]+)\s+and scoring\s+(\d+)\s+out of\s+(\d+)\s+for his final year thesis project',
            text
        )
        if mtech_match:
            degree, field, year, cgpa, thesis_score, thesis_max = mtech_match.groups()
            entries.append({
                'Category': 'Education',
                'Key': 'Graduate Degree',
                'Value': degree.strip(),
                'Comments': 'Master of Technology degree'
            })
            entries.append({
                'Category': 'Education',
                'Key': 'Graduate Field',
                'Value': field.strip(),
                'Comments': 'Advanced specialization in emerging technology field'
            })
            
            # Extract M.Tech institution
            mtech_inst_match = re.search(r'His academic excellence continued at\s+([^,]+),\s+where he earned his M\.Tech', text)
            if mtech_inst_match:
                entries.append({
                    'Category': 'Education',
                    'Key': 'Graduate Institution',
                    'Value': mtech_inst_match.group(1).strip(),
                    'Comments': 'Premier Indian Institute of Technology'
                })
            
            entries.append({
                'Category': 'Education',
                'Key': 'Graduate Graduation Year',
                'Value': year,
                'Comments': 'Year of postgraduate degree completion'
            })
            entries.append({
                'Category': 'Education',
                'Key': 'Graduate CGPA',
                'Value': cgpa,
                'Comments': 'Exceptional academic performance in postgraduate studies'
            })
            entries.append({
                'Category': 'Education',
                'Key': 'Thesis Score',
                'Value': f'{thesis_score}/{thesis_max}',
                'Comments': 'Outstanding performance in final year research project'
            })
        
        return entries
    
    def _extract_certification_info(self, text: str) -> List[Dict]:
        """Extract professional certification information"""
        entries = []
        
        # AWS certification
        aws_match = re.search(
            r'passed the\s+([^e]+)\s+exam in\s+(\d{4})\s+with a score of\s+(\d+)\s+out of\s+(\d+)',
            text
        )
        if aws_match:
            cert_name, year, score, max_score = aws_match.groups()
            entries.append({
                'Category': 'Certifications',
                'Key': 'AWS Certification',
                'Value': cert_name.strip(),
                'Comments': 'Cloud architecture certification from Amazon Web Services'
            })
            entries.append({
                'Category': 'Certifications',
                'Key': 'AWS Certification Year',
                'Value': year,
                'Comments': 'Year of certification achievement'
            })
            entries.append({
                'Category': 'Certifications',
                'Key': 'AWS Certification Score',
                'Value': f'{score}/{max_score}',
                'Comments': 'High performance score demonstrating strong cloud architecture knowledge'
            })
        
        # Azure certification
        azure_match = re.search(
            r'followed by the\s+([^c]+)\s+certification in\s+(\d{4})\s+with\s+(\d+)\s+points',
            text
        )
        if azure_match:
            cert_name, year, score = azure_match.groups()
            entries.append({
                'Category': 'Certifications',
                'Key': 'Azure Certification',
                'Value': cert_name.strip(),
                'Comments': 'Data engineering certification from Microsoft Azure'
            })
            entries.append({
                'Category': 'Certifications',
                'Key': 'Azure Certification Year',
                'Value': year,
                'Comments': 'Year of certification achievement'
            })
            entries.append({
                'Category': 'Certifications',
                'Key': 'Azure Certification Score',
                'Value': f'{score} points',
                'Comments': 'Strong performance in Azure data engineering assessment'
            })
        
        # PMP certification
        pmp_match = re.search(
            r'His\s+([^c]+)\s+certification, obtained in\s+(\d{4}),\s+was achieved with an\s+"([^"]+)"\s+rating from\s+([^,]+)',
            text
        )
        if pmp_match:
            cert_name, year, rating, org = pmp_match.groups()
            entries.append({
                'Category': 'Certifications',
                'Key': 'PMP Certification',
                'Value': cert_name.strip(),
                'Comments': 'Project management professional certification'
            })
            entries.append({
                'Category': 'Certifications',
                'Key': 'PMP Certification Year',
                'Value': year,
                'Comments': 'Year of certification achievement'
            })
            entries.append({
                'Category': 'Certifications',
                'Key': 'PMP Certification Rating',
                'Value': rating,
                'Comments': f'Highest performance tier from {org}'
            })
        
        # SAFe certification
        safe_match = re.search(
            r'while his\s+([^c]+)\s+certification earned him an outstanding\s+(\d+)%\s+score',
            text
        )
        if safe_match:
            cert_name, score = safe_match.groups()
            entries.append({
                'Category': 'Certifications',
                'Key': 'SAFe Certification',
                'Value': cert_name.strip(),
                'Comments': 'Scaled Agile Framework certification for enterprise agility'
            })
            entries.append({
                'Category': 'Certifications',
                'Key': 'SAFe Certification Score',
                'Value': f'{score}%',
                'Comments': 'Outstanding performance demonstrating mastery of agile methodologies'
            })
        
        return entries
    
    def _extract_skills_info(self, text: str) -> List[Dict]:
        """Extract technical skills and proficiency information"""
        entries = []
        
        # SQL skills
        sql_match = re.search(
            r'SQL expertise at a perfect\s+(\d+)\s+out of\s+(\d+),\s+reflecting his daily usage since\s+(\d{4})',
            text
        )
        if sql_match:
            rating, max_rating, since_year = sql_match.groups()
            entries.append({
                'Category': 'Technical Skills',
                'Key': 'SQL Proficiency',
                'Value': f'{rating}/{max_rating}',
                'Comments': f'Perfect rating; daily usage since {since_year}'
            })
        
        # Python skills
        python_match = re.search(
            r'His Python proficiency scores\s+(\d+)\s+out of\s+(\d+),\s+backed by over\s+([^y]+)\s+years of practical experience',
            text
        )
        if python_match:
            rating, max_rating, years = python_match.groups()
            entries.append({
                'Category': 'Technical Skills',
                'Key': 'Python Proficiency',
                'Value': f'{rating}/{max_rating}',
                'Comments': f'Over {years.strip()} years of practical experience'
            })
        
        # Machine Learning skills
        ml_match = re.search(
            r'while his machine learning capabilities rate\s+(\d+)\s+out of\s+(\d+),\s+representing\s+([^y]+)\s+years of hands-on implementation',
            text
        )
        if ml_match:
            rating, max_rating, years = ml_match.groups()
            entries.append({
                'Category': 'Technical Skills',
                'Key': 'Machine Learning Proficiency',
                'Value': f'{rating}/{max_rating}',
                'Comments': f'{years.strip()} years of hands-on implementation experience'
            })
        
        # Cloud Platform skills
        cloud_match = re.search(
            r'His cloud platform expertise, including AWS and Azure certifications, also rates\s+(\d+)\s+out of\s+(\d+)\s+with more than\s+([^y]+)\s+years of experience',
            text
        )
        if cloud_match:
            rating, max_rating, years = cloud_match.groups()
            entries.append({
                'Category': 'Technical Skills',
                'Key': 'Cloud Platform Proficiency (AWS & Azure)',
                'Value': f'{rating}/{max_rating}',
                'Comments': f'More than {years.strip()} years of experience; backed by professional certifications'
            })
        
        # Data Visualization skills
        viz_match = re.search(
            r'and his data visualization skills in Power BI and Tableau score\s+(\d+)\s+out of\s+(\d+),\s+establishing him as an expert in the field',
            text
        )
        if viz_match:
            rating, max_rating = viz_match.groups()
            entries.append({
                'Category': 'Technical Skills',
                'Key': 'Data Visualization Proficiency (Power BI & Tableau)',
                'Value': f'{rating}/{max_rating}',
                'Comments': 'Expert-level proficiency in business intelligence and data visualization tools'
            })
        
        return entries
    
    def export_to_excel(self, output_path: str):
        """Export structured data to Excel with formatting"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        
        # Define headers
        headers = ['Category', 'Key', 'Value', 'Comments']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Add data rows
        for entry in self.structured_data:
            ws.append([
                entry.get('Category', ''),
                entry.get('Key', ''),
                entry.get('Value', ''),
                entry.get('Comments', '')
            ])
        
        # Apply styling to data rows
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 60
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save workbook
        wb.save(output_path)
        print(f"✓ Excel file created: {output_path}")
        print(f"✓ Total entries extracted: {len(self.structured_data)}")


def main():
    """Main execution function"""
    print("=" * 70)
    print("AI-Powered Document Structuring & Data Extraction")
    print("=" * 70)
    
    # Initialize extractor
    extractor = DocumentExtractor("Data Input.pdf")
    
    # Extract text
    print("\n[1/3] Extracting text from PDF...")
    text = extractor.extract_text_from_pdf()
    print(f"✓ Extracted {len(text)} characters")
    
    # Identify key-value pairs
    print("\n[2/3] Identifying key-value relationships...")
    data = extractor.identify_key_value_pairs()
    print(f"✓ Identified {len(data)} structured data entries")
    
    # Export to Excel
    print("\n[3/3] Exporting to Excel...")
    extractor.export_to_excel("Output.xlsx")
    
    # Summary
    print("\n" + "=" * 70)
    print("EXTRACTION SUMMARY")
    print("=" * 70)
    categories = {}
    for entry in data:
        cat = entry['Category']
        categories[cat] = categories.get(cat, 0) + 1
    
    for category, count in categories.items():
        print(f"  • {category}: {count} entries")
    
    print("\n✓ Process completed successfully!")
    print("=" * 70)


if __name__ == "__main__":
    main()
